import datetime
import os
import sys

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import execute_values


# =========================================================
# CONFIGURACIÓN
# =========================================================

ARCHIVO_EXCEL = "alert_carga.xlsx"
HOJA_EXCEL = "CARGA MASIVA"

# Cambiar por el nombre real de la tabla PostgreSQL
TABLA_DESTINO = "adempiere.ofb_ssomac"

DB_CONFIG = {
    "host": os.getenv("PGHOST", "adempiere.tsm.cl"),
    "port": int(os.getenv("PGPORT", "5432")),
    "database": os.getenv("PGDATABASE", "tsm"),
    "user": os.getenv("PGUSER", "pg_api"),
    "password": os.getenv("PGPASSWORD", "8YR53mDRavJlfd6d"),
}


# Campos esperados en la hoja CARGA MASIVA
COLUMNAS_EXCEL = [
    "deviceentrytime",
    "datereport",
    "ad_user_id",
    "supervisor_id",
    "c_projectofb_id",
    "c_bpartner_id",
    "typealertcontrol",
    "a_asset_id",
    "documentno",
    "ofb_ssomac_id",
    "c_doctype_id",
]


# =========================================================
# CONVERSIÓN DE DATOS
# =========================================================

def convertir_fecha(valor, incluir_hora=True):
    """
    Convierte las fechas del Excel a objetos datetime.

    Formatos admitidos:
        01-05-2026
        01-05-2026 06:26:11
        2026-05-01
        2026-05-01 06:26:11
    """

    if valor is None or valor == "":
        return None

    if isinstance(valor, datetime.datetime):
        fecha = valor

    elif isinstance(valor, datetime.date):
        fecha = datetime.datetime.combine(
            valor,
            datetime.time.min
        )

    else:
        valor = str(valor).strip()

        formatos = [
            "%d-%m-%Y %H:%M:%S",
            "%d-%m-%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
        ]

        fecha = None

        for formato in formatos:
            try:
                fecha = datetime.datetime.strptime(valor, formato)
                break
            except ValueError:
                continue

        if fecha is None:
            raise ValueError(
                f"Formato de fecha no reconocido: {valor}"
            )

    if not incluir_hora:
        fecha = fecha.replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0
        )

    return fecha


def normalizar_entero(valor):
    """
    Convierte números provenientes de Excel a int.

    Si la celda viene vacía, retorna None para que PostgreSQL
    la guarde como NULL.
    """

    if valor is None or valor == "":
        return None

    return int(valor)


# =========================================================
# LECTURA DEL EXCEL
# =========================================================

def leer_excel(ruta_excel):
    if not os.path.exists(ruta_excel):
        raise FileNotFoundError(
            f"No se encontró el archivo: {ruta_excel}"
        )

    workbook = load_workbook(
        filename=ruta_excel,
        read_only=True,
        data_only=True
    )

    try:
        if HOJA_EXCEL not in workbook.sheetnames:
            raise ValueError(
                f"No existe la hoja '{HOJA_EXCEL}'. "
                f"Hojas disponibles: {workbook.sheetnames}"
            )

        hoja = workbook[HOJA_EXCEL]

        primera_fila = next(
            hoja.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            )
        )

        encabezados = [
            valor.strip() if isinstance(valor, str) else valor
            for valor in primera_fila
        ]

        columnas_faltantes = [
            columna
            for columna in COLUMNAS_EXCEL
            if columna not in encabezados
        ]

        if columnas_faltantes:
            raise ValueError(
                "Faltan columnas obligatorias en el Excel: "
                f"{columnas_faltantes}"
            )

        posiciones = {
            columna: encabezados.index(columna)
            for columna in COLUMNAS_EXCEL
        }

        fecha_actual = datetime.datetime.now()
        registros = []
        errores = []

        for numero_fila, fila in enumerate(
            hoja.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            # Ignorar filas completamente vacías
            if all(valor is None for valor in fila):
                continue

            try:
                datos = {
                    columna: fila[posiciones[columna]]
                    for columna in COLUMNAS_EXCEL
                }

                registro = (
                    # -----------------------------------------
                    # Campos nativos
                    # -----------------------------------------
                    1000000,       # ad_client_id
                    0,             # ad_org_id
                    fecha_actual,  # created
                    100,           # createdby
                    fecha_actual,  # updated
                    100,           # updatedby
                    "Y",           # isactive
                    "EO",          # docstatus

                    # -----------------------------------------
                    # Campos provenientes del Excel
                    # -----------------------------------------
                    convertir_fecha(
                        datos["deviceentrytime"],
                        incluir_hora=False
                    ),
                    convertir_fecha(
                        datos["datereport"],
                        incluir_hora=True
                    ),
                    normalizar_entero(datos["ad_user_id"]),
                    normalizar_entero(datos["supervisor_id"]),
                    normalizar_entero(datos["c_projectofb_id"]),
                    normalizar_entero(datos["c_bpartner_id"]),
                    normalizar_entero(datos["typealertcontrol"]),
                    normalizar_entero(datos["a_asset_id"]),
                    normalizar_entero(datos["documentno"]),
                    normalizar_entero(datos["ofb_ssomac_id"]),
                    normalizar_entero(datos["c_doctype_id"]),
                )

                registros.append(registro)

            except Exception as error:
                errores.append(
                    f"Fila {numero_fila}: {error}"
                )

        if errores:
            detalle = "\n".join(errores[:20])

            mensaje = (
                f"Se encontraron {len(errores)} filas "
                f"con errores:\n{detalle}"
            )

            if len(errores) > 20:
                mensaje += (
                    f"\n... y {len(errores) - 20} errores más."
                )

            raise ValueError(mensaje)

        return registros

    finally:
        workbook.close()


# =========================================================
# INSERCIÓN EN POSTGRESQL
# =========================================================

def insertar_registros(registros):
    if not registros:
        print("No existen registros para insertar.")
        return 0

    consulta = f"""
        INSERT INTO {TABLA_DESTINO} (
            ad_client_id,
            ad_org_id,
            created,
            createdby,
            updated,
            updatedby,
            isactive,
            docstatus,
            deviceentrytime,
            datereport,
            ad_user_id,
            supervisor_id,
            c_projectofb_id,
            c_bpartner_id,
            typealertcontrol,
            a_asset_id,
            documentno,
            ofb_ssomac_id,
            c_doctype_id
        )
        VALUES %s
    """

    conexion = None

    try:
        conexion = psycopg2.connect(**DB_CONFIG)

        with conexion.cursor() as cursor:
            execute_values(
                cursor,
                consulta,
                registros,
                page_size=500
            )

        conexion.commit()

        return len(registros)

    except Exception:
        if conexion:
            conexion.rollback()

        raise

    finally:
        if conexion:
            conexion.close()


# =========================================================
# PROCESO PRINCIPAL
# =========================================================

def ejecutar_carga(ruta_excel):
    print("========================================")
    print("CARGA MASIVA DE ALERTAS")
    print("========================================")
    print(f"Archivo: {ruta_excel}")
    print(f"Hoja: {HOJA_EXCEL}")
    print(f"Tabla: {TABLA_DESTINO}")
    print()

    print("Leyendo archivo Excel...")

    registros = leer_excel(ruta_excel)

    print(f"Registros encontrados: {len(registros)}")
    print("Insertando registros en PostgreSQL...")

    cantidad_insertada = insertar_registros(registros)

    print()
    print("========================================")
    print("CARGA FINALIZADA CORRECTAMENTE")
    print("========================================")
    print(f"Registros insertados: {cantidad_insertada}")


# =========================================================
# EJECUCIÓN DEL SCRIPT
# =========================================================

if __name__ == "__main__":
    # Permite entregar opcionalmente la ruta al ejecutar:
    # python cargar_alertas.py "archivo.xlsx"

    if len(sys.argv) >= 2:
        ARCHIVO_EXCEL = sys.argv[1]

    try:
        ejecutar_carga(ARCHIVO_EXCEL)

    except Exception as error:
        print()
        print("========================================")
        print("ERROR DURANTE LA CARGA")
        print("========================================")
        print(error)

        sys.exit(1)