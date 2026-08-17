import openpyxl
from fpdf import FPDF

def save_excel_sheet_as_pdf(excel_path, pdf_path, print_area="A1:R66"):
    # Cargar el libro de trabajo
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook[workbook.sheetnames[0]]

    # Determinar el rango de celdas
    start_cell, end_cell = print_area.split(":")
    start_row = worksheet[start_cell].row
    start_col = worksheet[start_cell].column
    end_row = worksheet[end_cell].row
    end_col = worksheet[end_cell].column

    # Crear un objeto PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    # Ajustar columnas para PDF
    col_widths = [30] * (end_col - start_col + 1)  # Ajusta el ancho de columna

    # Iterar sobre las filas y columnas del rango especificado
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for i, cell in enumerate(row):
            value = str(cell.value) if cell.value is not None else ""
            pdf.cell(col_widths[i], 10, value, border=1)  # Celda con bordes
        pdf.ln()  # Nueva línea

    # Guardar el PDF
    pdf.output(pdf_path)
    print(f"PDF generado en: {pdf_path}")


# Definir rutas
excel_path = 'C:/Users/lramirez/Github/Pytthon_Proyect/extracted_files/form_detail.xlsx'
pdf_path = f'C:/Users/lramirez/Github/Api/asset/pdf/tem.pdf'

# Llamar a la función
save_excel_sheet_as_pdf(excel_path, pdf_path)
